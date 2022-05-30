import re
import requests
from openpyxl import load_workbook
from time import time
from bs4 import BeautifulSoup

filename = 'colors.xlsx'
workbook = load_workbook(filename)
sheet = workbook.active
rows = sheet.max_row

try:
    for i in range (1, rows + 1):
        try: 
            cell = sheet.cell(row = i + 1, column = 1).value
            cell = re.sub("\s+", "-", cell.strip())
            cell = re.sub("([()])", "", cell)
        except AttributeError:
            print("Row out of range")
            workbook.save(filename)
            exit(0)    
        
        URL = "https://icolorpalette.com/color/{}".format(cell)    
        
        page = requests.get(URL)
        
        parsed = BeautifulSoup(page.content, "html.parser")

        div = parsed.find_all("div", class_ = "col-sm-12 py-4 col-md-6")

        header = parsed.find("h1")

        if not header.find("small"):
            for tag in div: 
                divWithSpan = tag.find_all("div", class_ = "col-12")
                for div in divWithSpan:
                    if "Hex" in div.text: 
                        hex_string = div.get_text()
                        hex_string_formatted = " ".join(hex_string.split()).replace("Hex", "").replace(":", "").replace("#", "").strip()
                        print(hex_string_formatted)
                        sheet.cell(row = i+1, column = 2).value = hex_string_formatted
                        print('Color {} done!'.format(i+1))
        else: 
            print('Incorrect color!')
except KeyboardInterrupt:
    workbook.save(filename)